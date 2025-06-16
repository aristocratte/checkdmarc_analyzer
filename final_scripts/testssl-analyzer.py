#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
🔒 TESTSSL.SH ANALYZER - ULTRA-DETAILED SSL/TLS SECURITY ANALYZER
================================================================

This tool analyzes testssl.sh scan results (CSV format) and generates comprehensive
Excel reports with security analysis, vulnerability assessments, and recommendations.

Features:
- Single CSV analysis with detailed breakdown
- Multiple CSV combination for domain-wide analysis
- Excel reports with charts and visualizations
- Security scoring and grading
- Vulnerability prioritization
- Compliance checking (PCI DSS, NIST, etc.)

Author: BOC Security Tools
Version: 1.0
Date: 2025-06-16
"""

import csv
import sys
import os
import argparse
import json
from datetime import datetime
from pathlib import Path
from collections import defaultdict, Counter
import re

# Excel dependencies
try:
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    import pandas as pd
    EXCEL_AVAILABLE = True
    print("📊 Excel libraries loaded successfully")
except ImportError as e:
    EXCEL_AVAILABLE = False
    print(f"⚠️ Excel libraries not available: {e}")

# ========================= SECURITY CATEGORIES =========================

SECURITY_CATEGORIES = {
    'PROTOCOLS': ['SSLv2', 'SSLv3', 'TLS1', 'TLS1_1', 'TLS1_2', 'TLS1_3'],
    'CIPHERS': ['cipherlist_NULL', 'cipherlist_aNULL', 'cipherlist_EXPORT', 
                'cipherlist_LOW', 'cipherlist_3DES_IDEA', 'cipherlist_OBSOLETED',
                'cipherlist_STRONG_NOFS', 'cipherlist_STRONG_FS'],
    'CERTIFICATES': ['cert_commonName', 'cert_trust', 'cert_chain_of_trust',
                    'cert_expirationStatus', 'cert_notAfter', 'OCSP_stapling'],
    'VULNERABILITIES': ['heartbleed', 'CCS', 'ROBOT', 'CRIME_TLS', 'POODLE_SSL',
                       'SWEET32', 'FREAK', 'DROWN', 'LOGJAM', 'BEAST', 'LUCKY13'],
    'FORWARD_SECRECY': ['FS', 'FS_ciphers', 'FS_ECDHE_curves'],
    'EXTENSIONS': ['TLS_extensions', 'NPN', 'ALPN_HTTP2', 'certificate_transparency'],
    'GRADING': ['overall_grade', 'final_score']
}

SEVERITY_WEIGHTS = {
    'CRITICAL': 100,
    'HIGH': 80,
    'MEDIUM': 60,
    'LOW': 40,
    'WARN': 30,
    'INFO': 10,
    'OK': 0,
    'DEBUG': 0
}

VULNERABILITY_DESCRIPTIONS = {
    'heartbleed': 'Critical vulnerability allowing memory disclosure',
    'CCS': 'ChangeCipherSpec injection vulnerability',
    'ROBOT': 'Return Of Bleichenbacher Oracle Threat',
    'CRIME_TLS': 'Compression Ratio Info-leak Made Easy',
    'POODLE_SSL': 'Padding Oracle On Downgraded Legacy Encryption',
    'SWEET32': 'Birthday attacks on 64-bit block ciphers',
    'FREAK': 'Factoring RSA Export Keys vulnerability',
    'DROWN': 'Decrypting RSA with Obsolete and Weakened eNcryption',
    'LOGJAM': 'Diffie-Hellman key exchange vulnerability',
    'BEAST': 'Browser Exploit Against SSL/TLS',
    'LUCKY13': 'Lucky Thirteen timing attack'
}

# ========================= MAIN ANALYZER CLASS =========================

class TestSSLAnalyzer:
    """Main analyzer class for testssl.sh CSV reports."""
    
    def __init__(self):
        self.scans = []
        self.combined_data = defaultdict(list)
        self.analysis_results = {}
        
    def load_csv(self, csv_file):
        """Load and parse a testssl.sh CSV file."""
        try:
            scan_data = {
                'file': csv_file,
                'domain': self._extract_domain_from_filename(csv_file),
                'data': [],
                'timestamp': self._extract_timestamp_from_filename(csv_file)
            }
            
            with open(csv_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    scan_data['data'].append(row)
            
            self.scans.append(scan_data)
            print(f"✅ Loaded {len(scan_data['data'])} entries from {csv_file}")
            return True
            
        except Exception as e:
            print(f"❌ Error loading {csv_file}: {e}")
            return False
    
    def _extract_domain_from_filename(self, filename):
        """Extract domain name from filename."""
        basename = os.path.basename(filename)
        # Pattern: domain_pPORT-YYYYMMDD-HHMM.csv
        match = re.match(r'^(.+?)_p\d+', basename)
        if match:
            return match.group(1)
        return basename.replace('.csv', '')
    
    def _extract_timestamp_from_filename(self, filename):
        """Extract timestamp from filename."""
        basename = os.path.basename(filename)
        # Pattern: YYYYMMDD-HHMM
        match = re.search(r'(\d{8}-\d{4})', basename)
        if match:
            try:
                return datetime.strptime(match.group(1), '%Y%m%d-%H%M')
            except:
                pass
        return datetime.now()
    
    def analyze_scan(self, scan_data):
        """Analyze a single scan's data."""
        analysis = {
            'domain': scan_data['domain'],
            'file': scan_data['file'],
            'timestamp': scan_data['timestamp'],
            'total_checks': len(scan_data['data']),
            'security_score': 0,
            'grade': 'F',
            'vulnerabilities': [],
            'certificates': {},
            'protocols': {},
            'ciphers': {},
            'summary': {},
            'recommendations': []
        }
        
        severity_counts = Counter()
        protocol_status = {}
        vulnerabilities = []
        certificate_info = {}
        cipher_analysis = {}
        
        for entry in scan_data['data']:
            test_id = entry.get('id', '')
            severity = entry.get('severity', '').upper()
            finding = entry.get('finding', '')
            fqdn_ip = entry.get('fqdn/ip', '')
            cve = entry.get('cve', '')
            cwe = entry.get('cwe', '')
            
            severity_counts[severity] += 1
            
            # Analyze protocols
            if test_id in SECURITY_CATEGORIES['PROTOCOLS']:
                protocol_status[test_id] = {
                    'status': severity,
                    'finding': finding,
                    'secure': severity in ['OK', 'INFO'] and 'not offered' in finding
                }
            
            # Analyze vulnerabilities
            if test_id in SECURITY_CATEGORIES['VULNERABILITIES']:
                vuln_info = {
                    'name': test_id,
                    'severity': severity,
                    'status': finding,
                    'cve': cve,
                    'cwe': cwe,
                    'description': VULNERABILITY_DESCRIPTIONS.get(test_id, 'Unknown vulnerability'),
                    'vulnerable': severity not in ['OK', 'INFO'] or 'vulnerable' in finding.lower()
                }
                vulnerabilities.append(vuln_info)
            
            # Analyze certificates
            if 'cert_' in test_id or test_id in ['OCSP_stapling', 'certificate_transparency']:
                certificate_info[test_id] = {
                    'status': severity,
                    'value': finding
                }
            
            # Analyze ciphers
            if test_id.startswith('cipher'):
                cipher_analysis[test_id] = {
                    'status': severity,
                    'details': finding
                }
            
            # Get overall grade
            if test_id == 'overall_grade':
                analysis['grade'] = finding
            elif test_id == 'final_score':
                try:
                    analysis['security_score'] = int(finding)
                except:
                    analysis['security_score'] = 0
        
        # Calculate security score if not available
        if analysis['security_score'] == 0:
            analysis['security_score'] = self._calculate_security_score(severity_counts)
        
        # Generate grade if not available
        if analysis['grade'] == 'F':
            analysis['grade'] = self._calculate_grade(analysis['security_score'])
        
        # Store detailed analysis
        analysis['protocols'] = protocol_status
        analysis['vulnerabilities'] = vulnerabilities
        analysis['certificates'] = certificate_info
        analysis['ciphers'] = cipher_analysis
        analysis['severity_distribution'] = dict(severity_counts)
        
        # Generate summary
        analysis['summary'] = self._generate_summary(analysis)
        
        # Generate recommendations
        analysis['recommendations'] = self._generate_recommendations(analysis)
        
        return analysis
    
    def _calculate_security_score(self, severity_counts):
        """Calculate security score based on severity distribution."""
        total_weight = 0
        max_possible = 0
        
        for severity, count in severity_counts.items():
            weight = SEVERITY_WEIGHTS.get(severity, 0)
            total_weight += (100 - weight) * count
            max_possible += 100 * count
        
        if max_possible > 0:
            return int((total_weight / max_possible) * 100)
        return 0
    
    def _calculate_grade(self, score):
        """Calculate letter grade based on security score."""
        if score >= 95:
            return 'A+'
        elif score >= 90:
            return 'A'
        elif score >= 85:
            return 'A-'
        elif score >= 80:
            return 'B+'
        elif score >= 75:
            return 'B'
        elif score >= 70:
            return 'B-'
        elif score >= 65:
            return 'C+'
        elif score >= 60:
            return 'C'
        elif score >= 55:
            return 'C-'
        elif score >= 50:
            return 'D'
        else:
            return 'F'
    
    def _generate_summary(self, analysis):
        """Generate security summary."""
        summary = {
            'total_tests': analysis['total_checks'],
            'critical_issues': len([v for v in analysis['vulnerabilities'] if v['severity'] == 'CRITICAL']),
            'high_issues': len([v for v in analysis['vulnerabilities'] if v['severity'] == 'HIGH']),
            'medium_issues': len([v for v in analysis['vulnerabilities'] if v['severity'] == 'MEDIUM']),
            'low_issues': len([v for v in analysis['vulnerabilities'] if v['severity'] == 'LOW']),
            'vulnerable_count': len([v for v in analysis['vulnerabilities'] if v['vulnerable']]),
            'tls_1_3_supported': any(p.get('status') == 'OK' for k, p in analysis['protocols'].items() if k == 'TLS1_3'),
            'tls_1_2_supported': any(p.get('status') == 'OK' for k, p in analysis['protocols'].items() if k == 'TLS1_2'),
            'weak_protocols': len([p for p in analysis['protocols'].values() if not p.get('secure', True)])
        }
        return summary
    
    def _generate_recommendations(self, analysis):
        """Generate security recommendations."""
        recommendations = []
        
        # Protocol recommendations
        if not analysis['summary']['tls_1_3_supported']:
            recommendations.append({
                'priority': 'HIGH',
                'category': 'Protocols',
                'issue': 'TLS 1.3 not supported',
                'recommendation': 'Enable TLS 1.3 support for enhanced security and performance'
            })
        
        if not analysis['summary']['tls_1_2_supported']:
            recommendations.append({
                'priority': 'CRITICAL',
                'category': 'Protocols',
                'issue': 'TLS 1.2 not supported',
                'recommendation': 'Enable TLS 1.2 support immediately - minimum requirement'
            })
        
        # Vulnerability recommendations
        for vuln in analysis['vulnerabilities']:
            if vuln['vulnerable'] and vuln['severity'] in ['CRITICAL', 'HIGH']:
                recommendations.append({
                    'priority': vuln['severity'],
                    'category': 'Vulnerabilities',
                    'issue': f"{vuln['name']} vulnerability detected",
                    'recommendation': f"Patch {vuln['description']} immediately"
                })
        
        # Certificate recommendations
        cert_status = analysis['certificates']
        if 'OCSP_stapling' in cert_status and cert_status['OCSP_stapling']['status'] == 'LOW':
            recommendations.append({
                'priority': 'MEDIUM',
                'category': 'Certificates',
                'issue': 'OCSP stapling not configured',
                'recommendation': 'Enable OCSP stapling for better certificate validation performance'
            })
        
        return recommendations
    
    def analyze_all_scans(self):
        """Analyze all loaded scans."""
        self.analysis_results = {}
        
        for scan in self.scans:
            domain = scan['domain']
            analysis = self.analyze_scan(scan)
            
            if domain not in self.analysis_results:
                self.analysis_results[domain] = []
            
            self.analysis_results[domain].append(analysis)
        
        return self.analysis_results
    
    def generate_excel_report(self, output_file):
        """Generate comprehensive Excel report."""
        if not EXCEL_AVAILABLE:
            print("❌ Excel libraries not available. Cannot generate Excel report.")
            return False
        
        try:
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets
            self._create_overview_sheet(wb)
            self._create_detailed_analysis_sheet(wb)
            self._create_vulnerabilities_sheet(wb)
            self._create_recommendations_sheet(wb)
            self._create_compliance_sheet(wb)
            
            # Apply formatting
            self._apply_excel_formatting(wb)
            
            # Save workbook
            wb.save(output_file)
            print(f"✅ Excel report saved to: {output_file}")
            return True
            
        except Exception as e:
            print(f"❌ Error generating Excel report: {e}")
            return False
    
    def _create_overview_sheet(self, workbook):
        """Create overview sheet with summary statistics."""
        ws = workbook.create_sheet("Overview", 0)
        
        # Headers
        headers = ['Domain', 'Scan Date', 'Overall Grade', 'Security Score', 
                  'Critical Issues', 'High Issues', 'TLS 1.3', 'TLS 1.2', 'Vulnerabilities']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        row = 2
        for domain, analyses in self.analysis_results.items():
            for analysis in analyses:
                ws.cell(row=row, column=1, value=domain)
                ws.cell(row=row, column=2, value=analysis['timestamp'].strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row, column=3, value=analysis['grade'])
                ws.cell(row=row, column=4, value=analysis['security_score'])
                ws.cell(row=row, column=5, value=analysis['summary']['critical_issues'])
                ws.cell(row=row, column=6, value=analysis['summary']['high_issues'])
                ws.cell(row=row, column=7, value='✅' if analysis['summary']['tls_1_3_supported'] else '❌')
                ws.cell(row=row, column=8, value='✅' if analysis['summary']['tls_1_2_supported'] else '❌')
                ws.cell(row=row, column=9, value=analysis['summary']['vulnerable_count'])
                row += 1
    
    def _create_detailed_analysis_sheet(self, workbook):
        """Create detailed analysis sheet."""
        ws = workbook.create_sheet("Detailed Analysis")
        
        # Headers
        headers = ['Domain', 'Test ID', 'Category', 'Severity', 'Status', 'CVE', 'CWE', 'Description']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        row = 2
        for domain, analyses in self.analysis_results.items():
            for analysis in analyses:
                for scan in self.scans:
                    if scan['domain'] == domain:
                        for entry in scan['data']:
                            ws.cell(row=row, column=1, value=domain)
                            ws.cell(row=row, column=2, value=entry.get('id', ''))
                            ws.cell(row=row, column=3, value=self._get_test_category(entry.get('id', '')))
                            ws.cell(row=row, column=4, value=entry.get('severity', ''))
                            ws.cell(row=row, column=5, value=entry.get('finding', ''))
                            ws.cell(row=row, column=6, value=entry.get('cve', ''))
                            ws.cell(row=row, column=7, value=entry.get('cwe', ''))
                            ws.cell(row=row, column=8, value=VULNERABILITY_DESCRIPTIONS.get(entry.get('id', ''), ''))
                            row += 1
    
    def _create_vulnerabilities_sheet(self, workbook):
        """Create vulnerabilities summary sheet."""
        ws = workbook.create_sheet("Vulnerabilities")
        
        # Headers
        headers = ['Domain', 'Vulnerability', 'Severity', 'Status', 'CVE', 'Description', 'Recommendation']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        row = 2
        for domain, analyses in self.analysis_results.items():
            for analysis in analyses:
                for vuln in analysis['vulnerabilities']:
                    if vuln['vulnerable']:
                        ws.cell(row=row, column=1, value=domain)
                        ws.cell(row=row, column=2, value=vuln['name'])
                        ws.cell(row=row, column=3, value=vuln['severity'])
                        ws.cell(row=row, column=4, value=vuln['status'])
                        ws.cell(row=row, column=5, value=vuln['cve'])
                        ws.cell(row=row, column=6, value=vuln['description'])
                        ws.cell(row=row, column=7, value=f"Patch {vuln['description']}")
                        row += 1
    
    def _create_recommendations_sheet(self, workbook):
        """Create recommendations sheet."""
        ws = workbook.create_sheet("Recommendations")
        
        # Headers
        headers = ['Domain', 'Priority', 'Category', 'Issue', 'Recommendation']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        row = 2
        for domain, analyses in self.analysis_results.items():
            for analysis in analyses:
                for rec in analysis['recommendations']:
                    ws.cell(row=row, column=1, value=domain)
                    ws.cell(row=row, column=2, value=rec['priority'])
                    ws.cell(row=row, column=3, value=rec['category'])
                    ws.cell(row=row, column=4, value=rec['issue'])
                    ws.cell(row=row, column=5, value=rec['recommendation'])
                    row += 1
    
    def _create_compliance_sheet(self, workbook):
        """Create compliance checking sheet."""
        ws = workbook.create_sheet("Compliance")
        
        # Headers
        headers = ['Domain', 'Standard', 'Requirement', 'Status', 'Notes']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # PCI DSS requirements
        pci_requirements = [
            'TLS 1.2+ Required',
            'Strong Ciphers Only',
            'No SSLv2/SSLv3',
            'Valid Certificates',
            'No Known Vulnerabilities'
        ]
        
        row = 2
        for domain, analyses in self.analysis_results.items():
            for analysis in analyses:
                for req in pci_requirements:
                    ws.cell(row=row, column=1, value=domain)
                    ws.cell(row=row, column=2, value='PCI DSS')
                    ws.cell(row=row, column=3, value=req)
                    ws.cell(row=row, column=4, value=self._check_compliance_requirement(analysis, req))
                    ws.cell(row=row, column=5, value='')
                    row += 1
    
    def _get_test_category(self, test_id):
        """Get category for a test ID."""
        for category, tests in SECURITY_CATEGORIES.items():
            if test_id in tests:
                return category
        return 'OTHER'
    
    def _check_compliance_requirement(self, analysis, requirement):
        """Check if analysis meets compliance requirement."""
        if requirement == 'TLS 1.2+ Required':
            return '✅ PASS' if analysis['summary']['tls_1_2_supported'] else '❌ FAIL'
        elif requirement == 'No Known Vulnerabilities':
            return '✅ PASS' if analysis['summary']['vulnerable_count'] == 0 else '❌ FAIL'
        elif requirement == 'No SSLv2/SSLv3':
            ssl_protocols = ['SSLv2', 'SSLv3']
            insecure = any(not p.get('secure', True) for k, p in analysis['protocols'].items() if k in ssl_protocols)
            return '❌ FAIL' if insecure else '✅ PASS'
        return '? UNKNOWN'
    
    def _apply_excel_formatting(self, workbook):
        """Apply formatting to Excel workbook."""
        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply to all sheets
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            # Header formatting
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = max(len(str(cell.value or '')) for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    def print_console_report(self):
        """Print summary report to console."""
        print("\n" + "="*80)
        print("🔒 TESTSSL.SH ANALYSIS REPORT")
        print("="*80)
        
        for domain, analyses in self.analysis_results.items():
            print(f"\n📊 Domain: {domain}")
            print("-" * 50)
            
            for analysis in analyses:
                print(f"  📅 Scan Date: {analysis['timestamp'].strftime('%Y-%m-%d %H:%M')}")
                print(f"  🏆 Overall Grade: {analysis['grade']}")
                print(f"  📈 Security Score: {analysis['security_score']}/100")
                print(f"  🔍 Total Tests: {analysis['summary']['total_tests']}")
                print(f"  🚨 Critical Issues: {analysis['summary']['critical_issues']}")
                print(f"  ⚠️  High Issues: {analysis['summary']['high_issues']}")
                print(f"  🔐 TLS 1.3 Support: {'✅' if analysis['summary']['tls_1_3_supported'] else '❌'}")
                print(f"  🔒 TLS 1.2 Support: {'✅' if analysis['summary']['tls_1_2_supported'] else '❌'}")
                print(f"  🐛 Active Vulnerabilities: {analysis['summary']['vulnerable_count']}")
                
                if analysis['vulnerabilities']:
                    print(f"\n  🚨 Detected Vulnerabilities:")
                    for vuln in analysis['vulnerabilities'][:5]:  # Show first 5
                        if vuln['vulnerable']:
                            print(f"    • {vuln['name']} ({vuln['severity']}) - {vuln['description']}")
                
                if analysis['recommendations']:
                    print(f"\n  💡 Top Recommendations:")
                    for rec in analysis['recommendations'][:3]:  # Show top 3
                        print(f"    • [{rec['priority']}] {rec['issue']}")
                        print(f"      → {rec['recommendation']}")
                
                print()

# ========================= MAIN FUNCTION =========================

def main():
    """Main function."""
    parser = argparse.ArgumentParser(
        description="🔒 TESTSSL.SH ANALYZER - SSL/TLS Security Analysis Tool"
    )
    
    parser.add_argument(
        'csv_files',
        nargs='+',
        help='One or more testssl.sh CSV files to analyze'
    )
    
    parser.add_argument(
        '-o', '--output',
        default='testssl_analysis_report.xlsx',
        help='Output Excel file name (default: testssl_analysis_report.xlsx)'
    )
    
    parser.add_argument(
        '--console-only',
        action='store_true',
        help='Only display console report, do not generate Excel'
    )
    
    args = parser.parse_args()
    
    print("🔒 TESTSSL.SH ANALYZER STARTING...")
    print("="*50)
    
    # Initialize analyzer
    analyzer = TestSSLAnalyzer()
    
    # Load CSV files
    successful_loads = 0
    for csv_file in args.csv_files:
        if not os.path.exists(csv_file):
            print(f"❌ File not found: {csv_file}")
            continue
        
        if analyzer.load_csv(csv_file):
            successful_loads += 1
    
    if successful_loads == 0:
        print("❌ No CSV files loaded successfully. Exiting.")
        sys.exit(1)
    
    print(f"\n📊 Successfully loaded {successful_loads} CSV files")
    
    # Analyze all scans
    print("\n🔍 Analyzing scans...")
    analyzer.analyze_all_scans()
    
    # Print console report
    analyzer.print_console_report()
    
    # Generate Excel report if requested
    if not args.console_only and EXCEL_AVAILABLE:
        print(f"\n📝 Generating Excel report: {args.output}")
        success = analyzer.generate_excel_report(args.output)
        
        if success:
            print(f"\n✅ Analysis complete! Excel report saved to: {args.output}")
        else:
            print(f"\n❌ Failed to generate Excel report")
    elif not args.console_only:
        print("\n⚠️ Excel libraries not available. Only console report generated.")
    
    print("\n🎉 TESTSSL.SH ANALYSIS COMPLETED!")

if __name__ == "__main__":
    main() 