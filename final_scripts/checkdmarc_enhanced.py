#!/usr/bin/env python3
"""
checkdmarc_enhanced.py
ULTRA-DETAILED analysis of a JSON report produced by checkdmarc
( SPF / DKIM / DMARC / MTA-STS / TLS-RPT / DNSSEC / BIMI )

🎯 This version EXPLAINS in detail:
- WHY each element is important
- WHAT HAPPENS if it is misconfigured
- WHAT ARE THE CONCRETE RISKS
- HOW TO FIX the issues

Usage:
    python3 checkdmarc_enhanced.py scan.json [scan2.json ...]

Output:
    - Exhaustive diagnostics with detailed explanations
    - Return code 0 if no CRITICAL, 1 otherwise (useful in CI/CD)

Author: ChatGPT & AI Assistant (2025)
"""
import json
import sys
from pathlib import Path
from typing import List, Dict, Tuple, Optional
import argparse
try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    EXCEL_AVAILABLE = True
    print("📊 Excel libraries loaded successfully")
except ImportError as e:
    EXCEL_AVAILABLE = False
    print(f"⚠️ Excel import error: {e}")
except Exception as e:
    EXCEL_AVAILABLE = False
    print(f"⚠️ Excel error: {e}")

# ----------------- Improved Official References -----------------

REF: Dict[str, str] = {
    "SPF_LIMIT": "RFC 7208 §4.6.4 – 10 DNS lookup limit | https://tools.ietf.org/html/rfc7208#section-4.6.4",
    "SPF_ALL": "NIST SP 800-177r1 §2.3 – Use of '-all' | https://csrc.nist.gov/publications/detail/sp/800-177/rev-1/final",
    "DKIM_LEN": "RFC 8301 – Minimum 1024 bits, recommended 2048 | https://tools.ietf.org/html/rfc8301",
    "DMARC_POLICY": "RFC 7489 §6.3 – p=none/quarantine/reject | https://tools.ietf.org/html/rfc7489#section-6.3",
    "MTA_STS": "RFC 8461 – MTA-STS enforce vs testing | https://tools.ietf.org/html/rfc8461",
    "TLS_RPT": "RFC 8460 – SMTP TLS Reporting | https://tools.ietf.org/html/rfc8460",
    "DNSSEC": "RFC 4033-35 – DNS Authentication | https://tools.ietf.org/html/rfc4033",
    "BIMI": "BIMI WG draft – DMARC p=quarantine/reject required | https://datatracker.ietf.org/doc/draft-brand-indicators-for-message-identification/"
}

# ----------------- Improved Helpers -----------------

Status = Tuple[str, str, str]  # (LEVEL, MESSAGE, REFKEY)

def status(level: str, msg: str, ref: str) -> Status:
    return (level, msg, ref)

def print_status(s: Status) -> None:
    lvl, msg, ref = s
    icons = {"OK": "✅", "WARNING": "⚠️", "CRITICAL": "🚨", "INFO": "ℹ️"}
    colors = {"OK": "\033[92m", "WARNING": "\033[93m", "CRITICAL": "\033[91m", "INFO": "\033[94m"}
    
    icon = icons.get(lvl, "❓")
    color = colors.get(lvl, "\033[0m")
    reset = "\033[0m"
    
    print(f"{icon} [{color}{lvl}{reset}] {msg}")
    print(f"   📚 Reference: {REF.get(ref, ref)}")
    print()

# ----------------- ULTRA-DETAILED SPF ANALYSIS -----------------

def analyze_spf(spf: dict) -> List[Status]:
    out: List[Status] = []
    
    if not spf or not spf.get("record"):
        out.append(status("CRITICAL",
                          "NO SPF RECORD FOUND! 🚨\n"
                          "💥 CRITICAL: No SPF protection in place\n"
                          "🎯 CONSEQUENCES:\n"
                          "   • Your domain can be spoofed by anyone\n"
                          "   • Major deliverability issues\n"
                          "   • Reputation and trust at risk\n"
                          "🛠️ FIX: Publish a valid SPF record in DNS\n"
                          "   Example: 'v=spf1 ip4:1.2.3.4 include:_spf.google.com -all'\n"
                          "📊 PRIORITY: URGENT (fix within hours)",
                          "SPF_LIMIT"))
        return out

    record = spf["record"]
    out.append(status("INFO", f"📝 SPF record detected: {record}", "SPF_LIMIT"))
    
    # Validity
    if not spf.get("valid", False):
        out.append(status("CRITICAL",
                          "INVALID SPF RECORD! 🚨\n"
                          "💥 The SPF record is present but contains syntax or logic errors.\n"
                          "🎯 CONSEQUENCES:\n"
                          "   • SPF will be ignored by recipient servers\n"
                          "   • No protection against spoofing\n"
                          "   • Deliverability issues likely\n"
                          "🛠️ FIX: Validate and correct the SPF record using online tools\n"
                          "   Example: https://dmarcian.com/spf-survey/",
                          "SPF_LIMIT"))
    else:
        out.append(status("OK",
                          "SPF record is syntactically valid.\n"
                          "✅ Syntax and logic checks passed.",
                          "SPF_LIMIT"))

    # DNS lookups analysis
    dns_lookups = spf.get("dns_lookups", 0)
    dns_void_lookups = spf.get("dns_void_lookups", 0)
    total_lookups = dns_lookups + dns_void_lookups
    
    if total_lookups > 10:
        out.append(status("CRITICAL",
                          f"SPF TOO COMPLEX - GUARANTEED FAILURE!\n"
                          f"🔥 PROBLEM: {total_lookups} DNS lookups (RFC limit: 10 maximum)\n"
                          f"   • Normal lookups: {dns_lookups}\n"
                          f"   • Void lookups: {dns_void_lookups}\n"
                          f"💥 DISASTROUS CONSEQUENCE: Servers return 'PermError' and COMPLETELY IGNORE your SPF!\n"
                          f"⚠️ COMMON CAUSES:\n"
                          f"   • Too many 'include:' statements\n"
                          f"   • Complex redirection chains\n"
                          f"   • Recursive inclusions\n"
                          f"🛠️ IMMEDIATE SOLUTIONS:\n"
                          f"   1. Replace 'include:' with direct IPs (ip4:/ip6:)\n"
                          f"   2. Remove unnecessary inclusions\n"
                          f"   3. Use subdomains to divide rules\n"
                          f"📊 BUSINESS IMPACT: Zero protection + degraded deliverability!",
                          "SPF_LIMIT"))
    elif total_lookups > 7:
        out.append(status("WARNING",
                          f"SPF CLOSE TO CRITICAL LIMIT!\n"
                          f"⚠️ STATUS: {total_lookups}/10 DNS lookups used\n"
                          f"   • Normal lookups: {dns_lookups}\n"
                          f"   • Void lookups: {dns_void_lookups}\n"
                          f"🎯 RISK: Limit exceeded with future additions\n"
                          f"📈 TREND: Natural growth with infrastructure evolution\n"
                          f"🔮 PREDICTION: SPF failure probable in 6-12 months\n"
                          f"💡 PREVENTIVE RECOMMENDATION:\n"
                          f"   • Optimize now (easier than emergency fixes)\n"
                          f"   • Document necessary inclusions\n"
                          f"   • Plan refactoring if > 8 lookups",
                          "SPF_LIMIT"))
    elif total_lookups > 5:
        out.append(status("INFO",
                          f"SPF moderate complexity ({total_lookups}/10 DNS lookups)\n"
                          f"   • Normal lookups: {dns_lookups}\n"
                          f"   • Void lookups: {dns_void_lookups}\n"
                          f"✅ STATUS: Functional and within standards\n"
                          f"🎯 ADVICE: Monitor evolution with future additions\n"
                          f"📋 MAINTENANCE: Review annually for optimization",
                          "SPF_LIMIT"))
    elif total_lookups > 0:
        out.append(status("OK",
                          f"SPF OPTIMIZED! ({total_lookups}/10 DNS lookups)\n"
                          f"   • Normal lookups: {dns_lookups}\n"
                          f"   • Void lookups: {dns_void_lookups}\n"
                          f"✅ PERFORMANCE: Excellent\n"
                          f"🎯 MARGIN: Large margin for future evolution\n"
                          f"🏆 STATUS: Optimal configuration",
                          "SPF_LIMIT"))

    # ALL directive analysis (most important!)
    if record and record.strip().endswith("-all"):
        out.append(status("OK", 
                          "MAXIMUM SPF PROTECTION ENABLED! 🛡️\n"
                          "🎯 DIRECTIVE '-all' (FAIL) = Strictest policy\n"
                          "✅ OPERATION:\n"
                          "   • Authorized emails: ACCEPTED normally\n"
                          "   • Unauthorized emails: REJECTED outright\n"
                          "🏆 BUSINESS ADVANTAGES:\n"
                          "   • Anti-spoofing protection: 95%+\n"
                          "   • Enhanced client trust\n"
                          "   • Domain reputation preserved\n"
                          "   • Maximum security compliance\n"
                          "📊 RESULT: Your domain is TRULY protected!",
                          "SPF_ALL"))
    elif record and "~all" in record:
        out.append(status("WARNING",
                          "PARTIAL SPF PROTECTION - MODERATE RISK ⚠️\n"
                          "🎯 DIRECTIVE '~all' (SOFTFAIL) = Permissive policy\n"
                          "⚡ RISKY OPERATION:\n"
                          "   • Authorized emails: ACCEPTED normally\n"
                          "   • Unauthorized emails: ACCEPTED but marked 'suspicious'\n"
                          "🚨 FREQUENT PROBLEMS:\n"
                          "   • Many servers IGNORE the marking\n"
                          "   • Spoofing still possible\n"
                          "   • False sense of security\n"
                          "📈 STRATEGIC RECOMMENDATION:\n"
                          "   1. Test '-all' mode on test domain\n"
                          "   2. Monitor DMARC reports for 2-4 weeks\n"
                          "   3. Switch to '-all' for real protection\n"
                          "🎯 GOAL: 95% protection instead of 60%",
                          "SPF_ALL"))
    elif record and "+all" in record:
        out.append(status("CRITICAL",
                          "SUICIDAL SPF DIRECTIVE DETECTED! 💀\n"
                          "🚨 DIRECTIVE '+all' (PASS) = SECURITY DISASTER\n"
                          "💥 CATASTROPHIC OPERATION:\n"
                          "   • ALL global servers authorized to send on your behalf\n"
                          "   • Spammers, hackers, competitors: free access\n"
                          "   • SPF turned into 'welcome fraudsters' sign\n"
                          "🔥 IMMEDIATE CONSEQUENCES:\n"
                          "   • Massive spoofing guaranteed\n"
                          "   • Reputation destroyed within hours\n"
                          "   • Blacklisting probable\n"
                          "   • Client trust loss\n"
                          "🆘 URGENT ACTION REQUIRED:\n"
                          "   REPLACE '+all' with '-all' IMMEDIATELY!\n"
                          "⏰ MAXIMUM DELAY: 1 hour (before malicious exploitation)",
                          "SPF_ALL"))
    elif record and "?all" in record:
        out.append(status("WARNING",
                          "SPF IN 'NEUTRAL' MODE - INEFFECTIVE! 🤷\n"
                          "🎯 DIRECTIVE '?all' (NEUTRAL) = No opinion\n"
                          "⚪ USELESS OPERATION:\n"
                          "   • SPF says 'I don't know' for unauthorized\n"
                          "   • Servers apply their local policy (unpredictable)\n"
                          "   • Variable behavior depending on recipients\n"
                          "📊 REAL PROTECTION: ~20% (random)\n"
                          "🎭 PROBLEM: False sense of security\n"
                          "🔧 SOLUTION: Choose '-all' or '~all' based on risk tolerance",
                          "SPF_ALL"))
    else:
        if record:  # SPF exists but no 'all' directive
            out.append(status("CRITICAL",
                              "INCOMPLETE SPF - NO 'ALL' DIRECTIVE! 🕳️\n"
                              "🚨 MAJOR PROBLEM: Truncated SPF record\n"
                              "⚡ UNPREDICTABLE BEHAVIOR:\n"
                              "   • Each server applies ITS default policy\n"
                              "   • Gmail: may accept or reject\n"
                              "   • Outlook: different behavior\n"
                              "   • Private servers: totally random\n"
                              "🎲 RESULT: Random protection (0-70%)\n"
                              "🎯 DIAGNOSIS: Configuration error or truncated record\n"
                              "🛠️ SIMPLE FIX: Add '-all' at end of record\n"
                              "📝 EXAMPLE: 'v=spf1 ip4:1.2.3.4 include:_spf.google.com -all'",
                              "SPF_ALL"))
    
    return out

# ----------------- ULTRA-DETAILED DKIM ANALYSIS -----------------

def analyze_dkim(dkim: dict) -> List[Status]:
    out: List[Status] = []
    
    if not dkim:
        out.append(status("WARNING",
                          "NO DKIM DETECTED IN SCAN! 🔍\n"
                          "⚠️ TECHNICAL LIMITATION: checkdmarc tests only one selector by default\n"
                          "🎯 STANDARD SELECTORS TESTED: 'default', 'selector1', 'dkim'\n"
                          "💡 POSSIBLE SITUATIONS:\n"
                          "   • DKIM exists but with a custom selector\n"
                          "   • Configured only on subdomains\n"
                          "   • DKIM keys being deployed\n"
                          "🔍 MANUAL VERIFICATION RECOMMENDED:\n"
                          "   1. Check outgoing email headers\n"
                          "   2. Look for 'DKIM-Signature:' in raw source\n"
                          "   3. Test custom selectors\n"
                          "📊 IMPACT: Incomplete authentication if truly absent",
                          "DKIM_LEN"))
        return out

    # Analysis of each DKIM selector
    for selector, det in dkim.items():
        out.append(status("INFO", f"🔑 Analyzing DKIM selector: '{selector}'", "DKIM_LEN"))
        
        record = det.get("record")
        if not record:
            out.append(status("CRITICAL",
                              f"DKIM SELECTOR '{selector}' TOTALLY ABSENT! 🚨\n"
                              f"💥 DIRECT CONSEQUENCE: DKIM signature impossible to verify\n"
                              f"⚡ AUTHENTICATION IMPACT:\n"
                              f"   • Emails marked 'DKIM=fail' or 'DKIM=none'\n"
                              f"   • DMARC cannot rely on DKIM\n"
                              f"   • Protection against transit modification = ZERO\n"
                              f"🎯 COMMON CAUSES:\n"
                              f"   • Key accidentally deleted from DNS\n"
                              f"   • Selector name error\n"
                              f"   • Incomplete DNS propagation\n"
                              f"🛠️ RESOLUTION:\n"
                              f"   1. Check mail server configuration\n"
                              f"   2. Regenerate DKIM key pair\n"
                              f"   3. Publish public key in DNS TXT\n"
                              f"📍 DNS Zone: {selector}._domainkey.yourdomain.com",
                              "DKIM_LEN"))
            continue

        if not det.get("valid", False):
            error_detail = det.get('error', 'unknown error')
            out.append(status("CRITICAL",
                              f"DKIM '{selector}' INVALID! 🔴\n"
                              f"💀 TECHNICAL ERROR: {error_detail}\n"
                              f"⚡ CONSEQUENCE: DKIM signature systematically rejected\n"
                              f"🎯 DELIVERABILITY IMPACT:\n"
                              f"   • Emails suspicious to recipient servers\n"
                              f"   • Degraded reputation score\n"
                              f"   • Increased risk of spam placement\n"
                              f"🔧 TYPICAL CAUSES:\n"
                              f"   • Incorrect public key format\n"
                              f"   • Invalid characters in record\n"
                              f"   • Corruption during DNS publication\n"
                              f"🛠️ DIAGNOSIS: Validate DNS record with DKIM tools",
                              "DKIM_LEN"))
            continue

        # Key strength analysis
        key_size = det.get("key_length", 0)
        if key_size < 1024:
            out.append(status("CRITICAL",
                              f"DKIM KEY '{selector}' DANGEROUSLY SHORT! ⚠️\n"
                              f"🔑 CURRENT SIZE: {key_size} bits (legal minimum: 1024 bits)\n"
                              f"💀 MAJOR CRYPTOGRAPHIC VULNERABILITY:\n"
                              f"   • Factorization possible in hours/days\n"
                              f"   • Attackers can forge your DKIM signatures\n"
                              f"   • Email spoofing with 'valid' authentication\n"
                              f"🚨 POSSIBLE EXPLOITATION:\n"
                              f"   • Undetectable phishing by filters\n"
                              f"   • Total authentication compromise\n"
                              f"⏰ IMMEDIATE ACTION REQUIRED:\n"
                              f"   1. Generate new 2048+ bit pair\n"
                              f"   2. Deploy new public key\n"
                              f"   3. Update server configuration\n"
                              f"   4. Test then remove old key\n"
                              f"📊 PRIORITY: CRITICAL (major security risk)",
                              "DKIM_LEN"))
        elif key_size < 2048:
            out.append(status("WARNING",
                              f"DKIM KEY '{selector}' SUB-OPTIMAL 📏\n"
                              f"🔑 CURRENT SIZE: {key_size} bits (recommended minimum: 2048 bits)\n"
                              f"⚠️ REDUCED SECURITY:\n"
                              f"   • Adequate protection today\n"
                              f"   • Increasing vulnerability over time\n"
                              f"   • Computing power constantly increasing\n"
                              f"🎯 STRATEGIC RECOMMENDATION:\n"
                              f"   • Plan migration to 2048 bits\n"
                              f"   • New industry standard\n"
                              f"   • Universal compatibility assured\n"
                              f"📅 SUGGESTED TIMEFRAME: 6-12 months (not urgent but recommended)\n"
                              f"🔐 2048 bits ADVANTAGE: 10+ years protection guaranteed",
                              "DKIM_LEN"))
        else:
            out.append(status("OK",
                              f"DKIM KEY '{selector}' EXCELLENT! 🏆\n"
                              f"🔑 SIZE: {key_size} bits (modern standard)\n"
                              f"✅ OPTIMAL CRYPTOGRAPHIC SECURITY:\n"
                              f"   • Protection against factorization: 10+ years\n"
                              f"   • Resistance to brute force attacks\n"
                              f"   • Compliance with current standards\n"
                              f"🎯 BUSINESS ADVANTAGES:\n"
                              f"   • Robust email authentication\n"
                              f"   • Guaranteed integrity in transit\n"
                              f"   • Maximum recipient trust\n"
                              f"   • Optimized deliverability\n"
                              f"🏅 RESULT: Exemplary DKIM configuration!",
                              "DKIM_LEN"))
    
    return out

# ----------------- ULTRA-DETAILED DMARC ANALYSIS -----------------

def analyze_dmarc(dmarc: dict) -> List[Status]:
    out: List[Status] = []
    
    if not dmarc or not dmarc.get("record"):
        out.append(status("CRITICAL",
                          "DMARC TOTALLY ABSENT! 🚨\n"
                          "💥 CRITICAL SITUATION: No anti-spoofing policy\n"
                          "🎯 DISASTROUS CONSEQUENCES:\n"
                          "   • SPF and DKIM exist but are USELESS\n"
                          "   • No instruction on what to do with failures\n"
                          "   • Servers apply random policies\n"
                          "   • Spoofing is possible even with SPF/DKIM\n"
                          "💀 MAJOR BUSINESS IMPACT:\n"
                          "   • Phishing using your domain\n"
                          "   • Reputation destroyed by third-party spam\n"
                          "   • Loss of client/partner trust\n"
                          "   • Legal and financial risks\n"
                          "🆘 IMMEDIATE FIX:\n"
                          "   Publish: 'v=DMARC1; p=none; rua=mailto:dmarc@yourdomain.com'\n"
                          "📊 PRIORITY: URGENT (fix in hours, not days)",
                          "DMARC_POLICY"))
        return out

    record = dmarc["record"]
    out.append(status("INFO", f"📋 DMARC policy detected: {record}", "DMARC_POLICY"))
    
    # Main policy analysis
    pvalue = dmarc["tags"]["p"]["value"]
    if pvalue == "none":
        out.append(status("CRITICAL",
                          "DMARC IN 'OBSERVATION' MODE ONLY! 👁️\n"
                          "⚠️ POLICY p=none = No active protection\n"
                          "📊 CURRENT OPERATION:\n"
                          "   • Fraudulent emails: ACCEPTED without restriction\n"
                          "   • Reports generated: YES (data collected)\n"
                          "   • Corrective action: NONE\n"
                          "🎯 LIMITED UTILITY:\n"
                          "   ✅ Monitoring and flow analysis\n"
                          "   ✅ Identification of missing legitimate sources\n"
                          "   ❌ Zero protection against spoofing\n"
                          "📈 RECOMMENDED PROGRESSION:\n"
                          "   1. Analyze DMARC reports for 4-6 weeks\n"
                          "   2. Identify missing legitimate sources\n"
                          "   3. Fix SPF/DKIM if necessary\n"
                          "   4. Move to p=quarantine then p=reject\n"
                          "⏰ GOAL: Active protection within 2-3 months maximum",
                          "DMARC_POLICY"))
    elif pvalue == "quarantine":
        out.append(status("WARNING",
                          "DMARC IN 'QUARANTINE' MODE - PARTIAL PROTECTION ⚠️\n"
                          "🎯 POLICY p=quarantine = Suspicious emails to spam\n"
                          "📊 CURRENT OPERATION:\n"
                          "   • Legitimate emails (SPF/DKIM OK): Inbox\n"
                          "   • Suspicious emails (auth failure): Spam/quarantine folder\n"
                          "   • Fraudulent emails: Generally blocked\n"
                          "✅ ADVANTAGES:\n"
                          "   • Active protection against 80-90% of attacks\n"
                          "   • Legitimate emails always delivered\n"
                          "   • Secure transition period\n"
                          "⚠️ LIMITATIONS:\n"
                          "   • Fraudulent emails sometimes visible (spam folder)\n"
                          "   • Users can access quarantines\n"
                          "   • Protection not absolute\n"
                          "🎯 STRATEGIC RECOMMENDATION:\n"
                          "   • Excellent intermediate step\n"
                          "   • Monitor reports for 4-8 weeks\n"
                          "   • Evolve to p=reject for maximum protection\n"
                          "📊 PROTECTION LEVEL: Very good (85-90%)",
                          "DMARC_POLICY"))
    elif pvalue == "reject":
        out.append(status("OK",
                          "DMARC IN 'REJECT' MODE - MAXIMUM PROTECTION! 🛡️\n"
                          "🏆 POLICY p=reject = Optimal configuration\n"
                          "✅ PERFECT OPERATION:\n"
                          "   • Legitimate emails (SPF/DKIM OK): Delivered normally\n"
                          "   • Fraudulent emails: REJECTED before reception\n"
                          "   • Spoofing: Impossible or nearly impossible\n"
                          "🎯 MAXIMUM BUSINESS PROTECTION:\n"
                          "   • Domain reputation preserved: 95%+\n"
                          "   • Client trust maintained\n"
                          "   • Phishing using your domain: blocked\n"
                          "   • Security compliance: excellent\n"
                          "💎 COMPETITIVE ADVANTAGES:\n"
                          "   • Brand protected against abuse\n"
                          "   • Security differentiation\n"
                          "   • Reduced security incidents\n"
                          "📊 PROTECTION LEVEL: Optimal (95-98%)\n"
                          "🏅 CONGRATULATIONS: Exemplary DMARC configuration!",
                          "DMARC_POLICY"))

    # Application percentage analysis
    pct = dmarc["tags"].get("pct", {}).get("value", 100)
    if pct < 100:
        out.append(status("WARNING",
                          f"DMARC PARTIALLY APPLIED! ⚠️\n"
                          f"📊 CURRENT PERCENTAGE: {pct}% of emails processed\n"
                          f"🎯 MEANING:\n"
                          f"   • {pct}% of emails: DMARC policy applied\n"
                          f"   • {100-pct}% of emails: no policy (like p=none)\n"
                          f"⚠️ PARTIAL DEPLOYMENT RISKS:\n"
                          f"   • Attackers can exploit the unprotected {100-pct}%\n"
                          f"   • Random and unpredictable protection\n"
                          f"   • False sense of security\n"
                          f"🎯 LEGITIMATE USE: Progressive transition to full protection\n"
                          f"📈 RECOMMENDATION:\n"
                          f"   1. If tests OK for several weeks: move to 100%\n"
                          f"   2. If recent deployment: monitor and increase gradually\n"
                          f"   3. Final goal: pct=100 for complete protection\n"
                          f"⏰ RECOMMENDED DELAY: 4-8 weeks maximum in partial mode",
                          "DMARC_POLICY"))
    else:
        out.append(status("OK",
                          "DMARC APPLIED AT 100%! ✅\n"
                          "🎯 COMPLETE COVERAGE: All your emails protected\n"
                          "🛡️ UNIFORM PROTECTION: No exploitable vulnerabilities\n"
                          "📊 RESULT: Maximum and predictable security",
                          "DMARC_POLICY"))

    # Aggregate reports analysis (RUA)
    rua_warnings = dmarc.get("warnings", [])
    has_rua_warning = any("rua tag" in warning for warning in rua_warnings)
    
    if has_rua_warning:
        out.append(status("WARNING",
                          "DMARC REPORTS NOT CONFIGURED! 📊\n"
                          "⚠️ PROBLEM: No 'rua' address specified\n"
                          "💀 CONSEQUENCE: You're flying blind!\n"
                          "🎯 MAJOR IMPACTS:\n"
                          "   • No visibility on spoofing attempts\n"
                          "   • Impossible to detect missing legitimate sources\n"
                          "   • No feedback on your policy effectiveness\n"
                          "   • Problem diagnosis: impossible\n"
                          "🔍 LOST DATA:\n"
                          "   • Volume of emails processed daily\n"
                          "   • Unauthorized sending sources\n"
                          "   • SPF/DKIM success rates\n"
                          "   • Geographic location of attacks\n"
                          "🛠️ IMMEDIATE SOLUTION:\n"
                          "   Add: rua=mailto:dmarc-reports@yourdomain.com\n"
                          "📈 BENEFIT: Complete visibility on email security",
                          "DMARC_POLICY"))
    else:
        out.append(status("OK",
                          "DMARC REPORTS CONFIGURED! 📊\n"
                          "✅ ACTIVE MONITORING: Data collected daily\n"
                          "🎯 OPERATIONAL ADVANTAGES:\n"
                          "   • Proactive detection of spoofing attempts\n"
                          "   • Monitoring of legitimate sending sources\n"
                          "   • Continuous configuration optimization\n"
                          "   • Evidence for security investigations\n"
                          "📊 RECOMMENDATION: Analyze reports monthly",
                          "DMARC_POLICY"))
    
    return out

# ----------------- ULTRA-DETAILED MTA-STS ANALYSIS -----------------

def analyze_mta_sts(mta: dict) -> List[Status]:
    out: List[Status] = []
    
    if not mta or not mta.get("valid", False):
        error_detail = mta.get('error', 'not deployed') if mta else 'not deployed'
        out.append(status("WARNING",
                          f"MTA-STS NOT DEPLOYED! 🔐\n"
                          f"📋 STATUS: {error_detail}\n"
                          f"⚠️ SECURITY IMPACT:\n"
                          f"   • No protection against forced TLS downgrade\n"
                          f"   • Vulnerable to man-in-the-middle attacks\n"
                          f"   • Email encryption is optional (not guaranteed)\n"
                          f"🎯 MTA-STS EXPLAINED:\n"
                          f"   • Forces servers to use TLS (encryption)\n"
                          f"   • Prevents downgrade to unencrypted connections\n"
                          f"   • Validates recipient server certificates\n"
                          f"💡 OPTIONAL BUT RECOMMENDED DEPLOYMENT:\n"
                          f"   1. Create policy file at https://mta-sts.yourdomain.com\n"
                          f"   2. Publish DNS record _mta-sts.yourdomain.com\n"
                          f"   3. Set mode to 'enforce' after testing\n"
                          f"📊 PRIORITY: Medium (enhanced security)",
                          "MTA_STS"))
    else:
        mode = mta.get("policy", {}).get("mode", "unknown")
        out.append(status("OK",
                          f"MTA-STS SUCCESSFULLY DEPLOYED! 🔐\n"
                          f"🛡️ CURRENT MODE: {mode}\n"
                          f"✅ TLS PROTECTION ENABLED:\n"
                          f"   • Encrypted connections required\n"
                          f"   • Prevention of downgrade attacks\n"
                          f"   • Server certificate validation\n"
                          f"🎯 SECURITY BENEFITS:\n"
                          f"   • Emails protected in transit\n"
                          f"   • Resistance to interception\n"
                          f"   • Compliance with modern standards\n"
                          f"📊 CONFIGURATION: Excellent (advanced standard)",
                          "MTA_STS"))
    
    return out

# ----------------- ULTRA-DETAILED TLS-RPT ANALYSIS -----------------

def analyze_tlsrpt(tls: dict) -> List[Status]:
    out: List[Status] = []
    
    if not tls or not tls.get("valid", False):
        out.append(status("WARNING",
                          "TLS-RPT NOT CONFIGURED! 📊\n"
                          "⚠️ NO TLS MONITORING:\n"
                          "   • No visibility on encryption failures\n"
                          "   • TLS issues not detected automatically\n"
                          "   • Downgrade attacks invisible\n"
                          "🎯 TLS-RPT EXPLAINED:\n"
                          "   • Automatic reports on TLS failures\n"
                          "   • Proactive detection of delivery issues\n"
                          "   • Monitoring of transport security\n"
                          "💡 DEPLOYMENT BENEFITS:\n"
                          "   • Quick diagnosis of deliverability issues\n"
                          "   • Detection of interception attempts\n"
                          "   • Optimization of TLS configuration\n"
                          "🛠️ SIMPLE DEPLOYMENT:\n"
                          "   DNS record: _smtp._tls.yourdomain.com\n"
                          "   Value: v=TLSRPTv1; rua=mailto:tls-reports@yourdomain.com\n"
                          "📊 PRIORITY: Low (advanced monitoring)",
                          "TLS_RPT"))
    else:
        out.append(status("OK",
                          "TLS-RPT CONFIGURED! 📊\n"
                          "✅ ACTIVE TLS MONITORING:\n"
                          "   • Continuous monitoring of encryption failures\n"
                          "   • Automatic detection of issues\n"
                          "   • Detailed reports on TLS connections\n"
                          "🎯 OPERATIONAL BENEFITS:\n"
                          "   • Proactive resolution of delivery issues\n"
                          "   • Visibility into infrastructure health\n"
                          "   • Continuous security improvement\n"
                          "📊 CONFIGURATION: Advanced (proactive monitoring)",
                          "TLS_RPT"))
    
    return out

# ----------------- ULTRA-DETAILED DNSSEC ANALYSIS -----------------

def analyze_dnssec(enabled: bool) -> List[Status]:
    if enabled:
        return [status("OK",
                      "DNSSEC ENABLED - MAXIMUM DNS PROTECTION! 🔐\n"
                      "✅ ENHANCED DNS SECURITY:\n"
                      "   • Cryptographic authentication of DNS responses\n"
                      "   • Protection against DNS cache poisoning\n"
                      "   • Integrity guaranteed for SPF/DKIM/DMARC records\n"
                      "🎯 CRITICAL ADVANTAGES:\n"
                      "   • DNS spoofing attacks: impossible\n"
                      "   • Malicious redirection: blocked\n"
                      "   • Absolute trust in DNS resolutions\n"
                      "🏆 BUSINESS IMPACT:\n"
                      "   • Ultra-secure email infrastructure\n"
                      "   • Protection against sophisticated attacks\n"
                      "   • Compliance with advanced security standards\n"
                      "📊 LEVEL: Security excellence (top 5% of domains)",
                      "DNSSEC")]
    
    return [status("WARNING",
                  "DNSSEC NOT DEPLOYED! 🔓\n"
                  "⚠️ DNS VULNERABILITY:\n"
                  "   • DNS responses not authenticated\n"
                  "   • Risk of DNS cache poisoning\n"
                  "   • Possibility of malicious redirection\n"
                  "🎯 POSSIBLE ATTACKS:\n"
                  "   • Hijacking of SPF/DKIM records\n"
                  "   • Redirecting emails to malicious servers\n"
                  "   • Compromising email authentication\n"
                  "💡 DNSSEC EXPLAINED:\n"
                  "   • Cryptographic signature of DNS zones\n"
                  "   • Authenticity validation by resolvers\n"
                  "   • Chain of trust from root servers\n"
                  "🛠️ DEPLOYMENT:\n"
                  "   • Contact registrar/DNS host\n"
                  "   • Usually free activation\n"
                  "   • Technical configuration required\n"
                  "📊 PRIORITY: Medium (enhanced security)",
                  "DNSSEC")]

# ----------------- ULTRA-DETAILED BIMI ANALYSIS -----------------

def analyze_bimi(bimi: dict, dmarc_policy: str) -> List[Status]:
    out: List[Status] = []
    
    if not bimi or not bimi.get("record"):
        out.append(status("INFO",
                          "BIMI NOT DEPLOYED (NORMAL) 🎨\n"
                          "📋 STATUS: Optional - Marketing impact only\n"
                          "🎯 BIMI EXPLAINED:\n"
                          "   • Brand Indicators for Message Identification\n"
                          "   • Displays your brand logo in email clients\n"
                          "   • Strengthens visual recognition of your emails\n"
                          "💡 MARKETING ADVANTAGES:\n"
                          "   • Improved brand recognition\n"
                          "   • Visual differentiation in inbox\n"
                          "   • Increased user trust\n"
                          "   • Reduced visual phishing\n"
                          "⚠️ STRICT PREREQUISITES:\n"
                          "   • DMARC with p=quarantine or p=reject REQUIRED\n"
                          "   • VMC (Verified Mark Certificate) required\n"
                          "   • Logo in specific SVG format\n"
                          "📊 PRIORITY: Very low (cosmetic/marketing)",
                          "BIMI"))
        return out

    if not bimi.get("valid", False):
        error_detail = bimi.get('error', 'invalid configuration')
        out.append(status("WARNING",
                          f"BIMI INVALID! 🎨\n"
                          f"🔴 ERROR: {error_detail}\n"
                          f"⚠️ CONSEQUENCE: Logo not displayed in email clients\n"
                          f"🎯 COMMON CAUSES:\n"
                          f"   • SVG format not compliant with specifications\n"
                          f"   • Logo URL inaccessible or incorrect\n"
                          f"   • VMC certificate missing or invalid\n"
                          f"   • Incorrect DNS record syntax\n"
                          f"🛠️ RECOMMENDED DIAGNOSIS:\n"
                          f"   1. Validate SVG format with BIMI tools\n"
                          f"   2. Check logo URL accessibility\n"
                          f"   3. Check VMC certificate validity\n"
                          f"📊 IMPACT: Cosmetic only (not security)",
                          "BIMI"))
    else:
        out.append(status("OK",
                          "BIMI SUCCESSFULLY CONFIGURED! 🎨\n"
                          "✅ BRAND LOGO ACTIVE:\n"
                          "   • Logo displayed in Gmail, Yahoo, etc.\n"
                          "   • Enhanced visual identity\n"
                          "   • Premium differentiation in inboxes\n"
                          "🎯 MARKETING BENEFITS ACHIEVED:\n"
                          "   • Immediate recognition of your emails\n"
                          "   • Increased recipient trust\n"
                          "   • Protection against visual spoofing\n"
                          "📊 STATUS: Optimal marketing configuration",
                          "BIMI"))

    # DMARC prerequisite check for BIMI
    if dmarc_policy not in ["reject", "quarantine"]:
        out.append(status("WARNING",
                          "BIMI WITHOUT STRICT DMARC! ⚠️\n"
                          f"🚨 PROBLEM: DMARC in mode '{dmarc_policy}' (required: quarantine/reject)\n"
                          "💀 CONSEQUENCE: BIMI logo ignored by most clients\n"
                          "🎯 TECHNICAL EXPLANATION:\n"
                          "   • BIMI requires strong anti-spoofing protection\n"
                          "   • Gmail/Yahoo refuse to display logos without strict DMARC\n"
                          "   • BIMI investment wasted without prior protection\n"
                          "🛠️ FIX:\n"
                          "   1. Change DMARC to p=quarantine or p=reject\n"
                          "   2. Wait for propagation (24-48h)\n"
                          "   3. BIMI will work automatically\n"
                          "📊 PRIORITY: Medium (fix DMARC first)",
                          "BIMI"))
    
    return out

# ----------------- ULTRA-DETAILED MX AND STARTTLS ANALYSIS -----------------

def analyze_mx_starttls(mx: dict) -> List[Status]:
    """Analyzes MX servers and their STARTTLS capabilities"""
    out: List[Status] = []
    
    if not mx or not mx.get("hosts"):
        out.append(status("CRITICAL",
                          "NO MX SERVERS CONFIGURED!\n"
                          "🚨 CRITICAL PROBLEM: Unable to receive emails\n"
                          "💥 IMMEDIATE CONSEQUENCES:\n"
                          "   • Incoming emails permanently lost\n"
                          "   • Client communications interrupted\n"
                          "   • Loss of business opportunities\n"
                          "   • Professional reputation damaged\n"
                          "🛠️ URGENT FIX:\n"
                          "   Configure at least one MX record\n"
                          "   Example: '10 mail.yourdomain.com'",
                          "SPF_LIMIT"))
        return out

    hosts = mx.get("hosts", [])
    out.append(status("INFO", f"📧 {len(hosts)} MX server(s) configured", "MTA_STS"))
    
    starttls_supported = 0
    starttls_failed = 0
    connection_issues = 0
    
    for i, host in enumerate(hosts):
        hostname = host.get("hostname", "unknown")
        preference = host.get("preference", 0)
        starttls = host.get("starttls", False)
        addresses = host.get("addresses", [])
        
        out.append(status("INFO", 
                          f"🖥️ MX Server #{i+1}: {hostname} (priority: {preference})\n"
                          f"   📍 IP Addresses: {', '.join(addresses) if addresses else 'Unresolved'}\n"
                          f"   🔐 STARTTLS: {'✅ Supported' if starttls else '❌ Not supported'}",
                          "MTA_STS"))
        
        if starttls:
            starttls_supported += 1
        else:
            starttls_failed += 1
    
    # Connection warnings
    warnings = mx.get("warnings", [])
    if warnings:
        connection_issues = len([w for w in warnings if "Connection" in w or "timed out" in w])
        out.append(status("WARNING",
                          f"CONNECTIVITY ISSUES DETECTED! ⚠️\n"
                          f"🚨 {len(warnings)} MX server(s) unreachable\n"
                          f"📝 Details:\n" + "\n".join([f"   • {w}" for w in warnings]) + "\n"
                          f"💡 POSSIBLE CAUSES:\n"
                          f"   • Servers temporarily offline\n"
                          f"   • Firewall blocking SMTP connections\n"
                          f"   • Incorrect DNS configuration\n"
                          f"   • Maintenance in progress\n"
                          f"🔧 RECOMMENDED ACTIONS:\n"
                          f"   1. Check server status with IT team\n"
                          f"   2. Test SMTP connectivity manually\n"
                          f"   3. Check firewall rules",
                          "MTA_STS"))

    # Global STARTTLS analysis
    if starttls_failed == 0 and starttls_supported > 0:
        out.append(status("OK",
                          "STARTTLS PERFECTLY CONFIGURED! 🔐\n"
                          f"✅ ALL MX servers ({starttls_supported}/{len(hosts)}) support STARTTLS\n"
                          "🛡️ OPTIMAL PROTECTION:\n"
                          "   • Incoming emails encrypted in transit\n"
                          "   • Protection against interception\n"
                          "   • Maximum security compliance\n"
                          "🏆 ACHIEVED BENEFITS:\n"
                          "   • Communication confidentiality\n"
                          "   • Regulatory compliance (GDPR, etc.)\n"
                          "   • Increased partner trust",
                          "MTA_STS"))
    elif starttls_supported > 0:
        out.append(status("WARNING",
                          f"STARTTLS PARTIALLY SUPPORTED! ⚠️\n"
                          f"📊 STATUS: {starttls_supported}/{len(hosts)} servers support STARTTLS\n"
                          f"🚨 RISK: Unencrypted emails on some servers\n"
                          f"💡 PROBLEM: Heterogeneous configuration\n"
                          f"🎯 SECURITY IMPACT:\n"
                          f"   • Potential interception vulnerability\n"
                          f"   • Partial non-compliance\n"
                          f"   • Risk depends on server used\n"
                          f"🛠️ FIX:\n"
                          f"   Enable STARTTLS on ALL MX servers",
                          "MTA_STS"))
    else:
        out.append(status("CRITICAL",
                          "NO MX SERVER SUPPORTS STARTTLS! 🚨\n"
                          "💥 ALL incoming emails are unencrypted in transit\n"
                          "⚠️ MAJOR RISK: Susceptible to interception and eavesdropping\n"
                          "🛠️ FIX: Enable STARTTLS on all MX servers as soon as possible",
                          "MTA_STS"))
    
    return out

# ----------------- SPECIFIC SECURITY CRITERIA ANALYSIS -----------------

def analyze_security_criteria(report: dict) -> List[Status]:
    """Checks the 10 specific security criteria from criteria.txt"""
    out: List[Status] = []
    
    out.append(status("INFO", "📋 CHECKING SPECIFIC SECURITY CRITERIA", "SPF_ALL"))
    
    # 1. SPF - SPF record present
    spf = report.get("spf", {})
    if spf.get("record") and spf.get("valid", False):
        out.append(status("OK", "SPF record present and valid.", "SPF_ALL"))
    else:
        out.append(status("CRITICAL", "SPF record missing or invalid!", "SPF_ALL"))
    
    # 2. SPF - Strict mode (checks if -all is used)
    spf_record = spf.get("record", "") or ""
    if spf_record.strip().endswith("-all"):
        out.append(status("OK", "SPF strict mode (-all) is enabled.", "SPF_ALL"))
    else:
        out.append(status("WARNING", "SPF strict mode (-all) is not enabled.", "SPF_ALL"))
    
    # 3. DMARC - DMARC record present
    dmarc = report.get("dmarc", {})
    if dmarc.get("record") and dmarc.get("valid", False):
        out.append(status("OK", "DMARC record present and valid.", "DMARC_POLICY"))
    else:
        out.append(status("CRITICAL", "DMARC record missing or invalid!", "DMARC_POLICY"))
    
    # 4. DMARC - Policy is not none
    dmarc_policy = dmarc.get("tags", {}).get("p", {}).get("value", "none")
    if dmarc_policy in ["quarantine", "reject"]:
        out.append(status("OK", f"✅ CRITERION 4/10: Strict DMARC policy (p={dmarc_policy})", "DMARC_POLICY"))
    else:
        out.append(status("CRITICAL", f"❌ CRITERION 4/10: Non-strict DMARC policy (p={dmarc_policy})", "DMARC_POLICY"))
    
    # 5. DMARC - Strict mode (checks if p=reject)
    if dmarc_policy == "reject":
        out.append(status("OK", "✅ CRITERION 5/10: DMARC in maximum strict mode (p=reject)", "DMARC_POLICY"))
    elif dmarc_policy == "quarantine":
        out.append(status("WARNING", "⚠️ CRITERION 5/10: DMARC moderately strict (p=quarantine, recommended: p=reject)", "DMARC_POLICY"))
    else:
        out.append(status("CRITICAL", f"❌ CRITERION 5/10: DMARC not in strict mode (p={dmarc_policy})", "DMARC_POLICY"))
    
    # 6. DMARC - rua present (aggregate reports)
    if "rua" in dmarc.get("tags", {}):
        out.append(status("OK", "✅ CRITERION 6/10: RUA address (aggregate reports) configured", "DMARC_POLICY"))
    else:
        out.append(status("CRITICAL", "❌ CRITERION 6/10: RUA address (aggregate reports) missing", "DMARC_POLICY"))
    
    # 7. DMARC - ruf present (detailed reports)
    if "ruf" in dmarc.get("tags", {}):
        out.append(status("OK", "✅ CRITERION 7/10: RUF address (detailed reports) configured", "DMARC_POLICY"))
    else:
        out.append(status("WARNING", "⚠️ CRITERION 7/10: RUF address (detailed reports) missing", "DMARC_POLICY"))
    
    # 8. DMARC - pct equals 100
    dmarc_pct = dmarc.get("tags", {}).get("pct", {}).get("value", 0)
    if dmarc_pct == 100:
        out.append(status("OK", "✅ CRITERION 8/10: DMARC applied to 100% of traffic (pct=100)", "DMARC_POLICY"))
    else:
        out.append(status("WARNING", f"⚠️ CRITERION 8/10: Partial DMARC (pct={dmarc_pct}%, recommended: 100%)", "DMARC_POLICY"))
    
    # 9. Mail Server - smtp - starttls offered
    mx = report.get("mx", {})
    mx_hosts = mx.get("hosts", [])
    starttls_count = sum(1 for host in mx_hosts if host.get("starttls", False))
    if starttls_count > 0 and starttls_count == len(mx_hosts):
        out.append(status("OK", f"✅ CRITERION 9/10: STARTTLS supported on all MX servers ({starttls_count}/{len(mx_hosts)})", "MTA_STS"))
    elif starttls_count > 0:
        out.append(status("WARNING", f"⚠️ CRITERION 9/10: Partial STARTTLS ({starttls_count}/{len(mx_hosts)} servers)", "MTA_STS"))
    else:
        out.append(status("CRITICAL", "❌ CRITERION 9/10: No MX server supports STARTTLS", "MTA_STS"))
    
    # 10. Mail Server - no pop service (this criterion requires external analysis)
    # Note: This information is not available in standard checkdmarc scan
    out.append(status("INFO", "ℹ️ CRITERION 10/10: POP service (requires manual verification)", "MTA_STS"))
    
    return out

# ----------------- Complete audit with explanations -----------------

def audit_domain(report: dict) -> List[Status]:
    results: List[Status] = []
    
    
    results += analyze_spf(report.get("spf"))
    
    
    results += analyze_dkim(report.get("dkim"))
    
    
    results += analyze_dmarc(report.get("dmarc"))
    
    
    results += analyze_mx_starttls(report.get("mx"))
    
    
    results += analyze_mta_sts(report.get("mta_sts"))
    
   
    results += analyze_tlsrpt(report.get("smtp_tls_reporting"))
    
    
    results += analyze_dnssec(report.get("dnssec", False))
    
    
    dmarc_policy = report.get("dmarc", {}).get("tags", {}).get("p", {}).get("value", "none")
    results += analyze_bimi(report.get("bimi"), dmarc_policy)
    
   
    results += analyze_security_criteria(report)
    
    return results

# ----------------- Excel Export Functions -----------------

def generate_excel_report(json_files: List[str], output_dir: Path) -> None:
    """
    Generates a complete Excel report with charts and detailed analysis.
    """
    if not EXCEL_AVAILABLE:
        print("❌ Excel libraries are not installed.")
        print("🔧 To install: pip install pandas openpyxl")
        return
    
    print("📊 Generating Excel report...")
    
    # Prepare data
    domains_data = []
    detailed_issues = []
    security_overview = {
        'SPF': {'OK': 0, 'WARNING': 0, 'CRITICAL': 0},
        'DKIM': {'OK': 0, 'WARNING': 0, 'CRITICAL': 0},
        'DMARC': {'OK': 0, 'WARNING': 0, 'CRITICAL': 0},
        'MTA-STS': {'OK': 0, 'WARNING': 0, 'CRITICAL': 0},
        'TLS-RPT': {'OK': 0, 'WARNING': 0, 'CRITICAL': 0},
        'DNSSEC': {'OK': 0, 'WARNING': 0, 'CRITICAL': 0},
        'BIMI': {'OK': 0, 'WARNING': 0, 'CRITICAL': 0}
    }
    
    # Process each JSON file
    for json_file in json_files:
        path = Path(json_file)
        if not path.exists():
            continue
            
        try:
            data = json.loads(path.read_text())
            domain_name = path.stem
            
            # Analyze domain
            statuses = audit_domain(data)
            
            # Count issues by category
            spf_status = get_component_status(statuses, 'SPF')
            dkim_status = get_component_status(statuses, 'DKIM')
            dmarc_status = get_component_status(statuses, 'DMARC')
            mta_sts_status = get_component_status(statuses, 'MTA-STS')
            tls_rpt_status = get_component_status(statuses, 'TLS-RPT')
            dnssec_status = get_component_status(statuses, 'DNSSEC')
            bimi_status = get_component_status(statuses, 'BIMI')
            
            # Update global statistics
            security_overview['SPF'][spf_status] += 1
            security_overview['DKIM'][dkim_status] += 1
            security_overview['DMARC'][dmarc_status] += 1
            security_overview['MTA-STS'][mta_sts_status] += 1
            security_overview['TLS-RPT'][tls_rpt_status] += 1
            security_overview['DNSSEC'][dnssec_status] += 1
            security_overview['BIMI'][bimi_status] += 1
            
            # Calculate global security score
            security_score = calculate_security_score(statuses)
            
            # Data for main table
            domain_info = {
                'Domain': domain_name,
                'Security Score': f"{security_score}%",
                'SPF': spf_status,
                'DKIM': dkim_status, 
                'DMARC': dmarc_status,
                'MTA-STS': mta_sts_status,
                'TLS-RPT': tls_rpt_status,
                'DNSSEC': dnssec_status,
                'BIMI': bimi_status,
                'Critical': sum(1 for s in statuses if s[0] == 'CRITICAL'),
                'Warnings': sum(1 for s in statuses if s[0] == 'WARNING'),
                'Global Status': get_overall_status(statuses)
            }
            domains_data.append(domain_info)
            
            # Issue details
            for status in statuses:
                if status[0] in ['CRITICAL', 'WARNING']:
                    issue_detail = {
                        'Domain': domain_name,
                        'Component': extract_component_from_message(status[1]),
                        'Severity': status[0],
                        'Description': clean_message_for_excel(status[1]),
                        'Reference': status[2]
                    }
                    detailed_issues.append(issue_detail)
                    
        except Exception as e:
            print(f"⚠️ Error processing {json_file}: {e}")
            continue
    
    # Create Excel file
    excel_file = output_dir / 'checkdmarc_security_report.xlsx'
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # Sheet 1: Overview
        df_overview = pd.DataFrame(domains_data)
        df_overview.to_excel(writer, sheet_name='Overview', index=False)
        
        # Sheet 2: Detailed issues
        df_issues = pd.DataFrame(detailed_issues)
        df_issues.to_excel(writer, sheet_name='Detailed Issues', index=False)
        
        # Sheet 3: Component statistics
        stats_data = []
        for component, stats in security_overview.items():
            total = sum(stats.values())
            if total > 0:
                stats_data.append({
                    'Component': component,
                    'Total Domains': total,
                    'OK': stats['OK'],
                    'Warnings': stats['WARNING'],
                    'Critical': stats['CRITICAL'],
                    '% OK': round((stats['OK'] / total) * 100, 1),
                    '% Issues': round(((stats['WARNING'] + stats['CRITICAL']) / total) * 100, 1)
                })
        
        df_stats = pd.DataFrame(stats_data)
        df_stats.to_excel(writer, sheet_name='Statistics', index=False)
    
    # Enhance formatting and add charts
    enhance_excel_formatting(excel_file, security_overview, domains_data)
    
    print(f"✅ Excel report generated successfully: {excel_file}")
    print("📊 The report contains:")
    print("   • Overview with security scores")
    print("   • Detailed issue analysis")
    print("   • Component statistics")
    print("   • Interactive charts")

def get_component_status(statuses: List[Status], component: str) -> str:
    """Determines component status based on analysis messages."""
    critical_found = False
    warning_found = False
    
    for status in statuses:
        message = status[1].upper()
        if component.upper() in message:
            if status[0] == 'CRITICAL':
                critical_found = True
            elif status[0] == 'WARNING':
                warning_found = True
    
    if critical_found:
        return 'CRITICAL'
    elif warning_found:
        return 'WARNING'
    else:
        return 'OK'

def calculate_security_score(statuses: List[Status]) -> int:
    """Calculates a global security score based on statuses."""
    total_checks = len(statuses)
    if total_checks == 0:
        return 0
    
    critical_count = sum(1 for s in statuses if s[0] == 'CRITICAL')
    warning_count = sum(1 for s in statuses if s[0] == 'WARNING')
    ok_count = sum(1 for s in statuses if s[0] == 'OK')
    
    # Weighted score: OK=100%, WARNING=50%, CRITICAL=0%
    score = (ok_count * 100 + warning_count * 50 + critical_count * 0) / total_checks
    return round(score)

def get_overall_status(statuses: List[Status]) -> str:
    """Determines the global domain status."""
    critical_count = sum(1 for s in statuses if s[0] == 'CRITICAL')
    warning_count = sum(1 for s in statuses if s[0] == 'WARNING')
    
    if critical_count > 0:
        return '🚨 CRITICAL'
    elif warning_count > 0:
        return '⚠️ WARNING'
    else:
        return '✅ EXCELLENT'

def extract_component_from_message(message: str) -> str:
    """Extracts the main component from error message."""
    message_upper = message.upper()
    components = ['SPF', 'DKIM', 'DMARC', 'MTA-STS', 'TLS-RPT', 'DNSSEC', 'BIMI']
    
    for comp in components:
        if comp in message_upper:
            return comp
    return 'GENERAL'

def clean_message_for_excel(message: str) -> str:
    """Clean message for Excel display."""
    # Remove color codes and emojis for Excel
    import re
    # Remove ANSI codes
    ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
    cleaned = ansi_escape.sub('', message)
    
    # Replace emojis with words
    emoji_replacements = {
        '🔥': '[DANGER]',
        '💥': '[IMPACT]',
        '🛠️': '[SOLUTION]',
        '⚡': '[RISKS]',
        '🔧': '[CORRECTION]',
        '✅': '[OK]',
        '🎯': '[RESULT]',
        '🛡️': '[PROTECTION]',
        '🚨': '[CRITICAL]',
        '⚠️': '[WARNING]',
        '🔴': '[ERROR]',
        '💀': '[CRITICAL IMPACT]'
    }
    
    for emoji, replacement in emoji_replacements.items():
        cleaned = cleaned.replace(emoji, replacement)
    
    # Limit length for Excel
    if len(cleaned) > 300:
        cleaned = cleaned[:297] + '...'
    
    return cleaned.strip()

def enhance_excel_formatting(excel_file: Path, security_overview: Dict, domains_data: List[Dict]) -> None:
    """Enhance Excel formatting and add charts."""
    wb = openpyxl.load_workbook(excel_file)
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    
    critical_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    ok_fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format the Overview sheet
    if 'Overview' in wb.sheetnames:
        ws_overview = wb['Overview']
        
        # Headers
        for cell in ws_overview[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data with conditional colors
        for row in ws_overview.iter_rows(min_row=2, max_row=ws_overview.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Coloring based on status
                if cell.value in ['CRITICAL', '🚨 CRITICAL']:
                    cell.fill = critical_fill
                elif cell.value in ['WARNING', '⚠️ WARNING']:
                    cell.fill = warning_fill
                elif cell.value in ['OK', '✅ EXCELLENT']:
                    cell.fill = ok_fill
        
        # Adjust column widths
        for column in ws_overview.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            ws_overview.column_dimensions[column[0].column_letter].width = min(max_length + 2, 30)
        
        # Add a pie chart for global status
        if len(domains_data) > 0:
            add_pie_chart_overview(ws_overview, domains_data)
    
    # Format the Statistics sheet
    if 'Statistics' in wb.sheetnames:
        ws_stats = wb['Statistics']
        
        # Headers
        for cell in ws_stats[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data
        for row in ws_stats.iter_rows(min_row=2, max_row=ws_stats.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adjust widths
        for column in ws_stats.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            ws_stats.column_dimensions[column[0].column_letter].width = min(max_length + 2, 20)
        
        # Add a bar chart
        add_bar_chart_stats(ws_stats, security_overview)
    
    # Format the Detailed Issues sheet
    if 'Detailed Issues' in wb.sheetnames:
        ws_issues = wb['Detailed Issues']
        
        # Headers
        for cell in ws_issues[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data with conditional colors
        for row in ws_issues.iter_rows(min_row=2, max_row=ws_issues.max_row):
            severity_cell = row[2]  # Severity column
            for cell in row:
                cell.border = border
                if cell.column_letter in ['A', 'B', 'C', 'E']:  # Specific columns
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                # Coloring based on severity
                if severity_cell.value == 'CRITICAL':
                    cell.fill = critical_fill
                elif severity_cell.value == 'WARNING':
                    cell.fill = warning_fill
        
        # Adjust widths
        ws_issues.column_dimensions['A'].width = 20  # Domain
        ws_issues.column_dimensions['B'].width = 15  # Component
        ws_issues.column_dimensions['C'].width = 12  # Severity
        ws_issues.column_dimensions['D'].width = 50  # Description
        ws_issues.column_dimensions['E'].width = 25  # Reference
        
        # Row height for text wrap
        for row_num in range(2, ws_issues.max_row + 1):
            ws_issues.row_dimensions[row_num].height = 60
    
    wb.save(excel_file)

def add_pie_chart_overview(worksheet, domains_data: List[Dict]) -> None:
    """Add a pie chart for global status."""
    # Count statuses
    status_counts = {}
    for domain in domains_data:
        status = domain['Global Status']
        status_counts[status] = status_counts.get(status, 0) + 1
    
    if not status_counts:
        return
    
    # Create data for the chart
    start_row = len(domains_data) + 5
    worksheet[f'A{start_row}'] = 'Status'
    worksheet[f'B{start_row}'] = 'Count'
    
    row = start_row + 1
    for status, count in status_counts.items():
        worksheet[f'A{row}'] = status
        worksheet[f'B{row}'] = count
        row += 1    
    # Create the chart
    chart = PieChart()
    chart.title = "Security Status Distribution"
    
    data = Reference(worksheet, min_col=2, min_row=start_row, max_row=row-1)
    labels = Reference(worksheet, min_col=1, min_row=start_row+1, max_row=row-1)
    
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)
    
    # Positioning
    chart.anchor = f'F{start_row}'
    chart.width = 15
    chart.height = 10
    
    worksheet.add_chart(chart)

def add_bar_chart_stats(worksheet, security_overview: Dict) -> None:
    """Add a bar chart for component statistics."""
    if not security_overview:
        return
    
    # Find a free position
    start_row = worksheet.max_row + 3
    
    # Create data for the chart
    worksheet[f'A{start_row}'] = 'Component'
    worksheet[f'B{start_row}'] = 'OK'
    worksheet[f'C{start_row}'] = 'Warnings'
    worksheet[f'D{start_row}'] = 'Critical'
    
    row = start_row + 1
    for component, stats in security_overview.items():
        total = sum(stats.values())
        if total > 0:
            worksheet[f'A{row}'] = component
            worksheet[f'B{row}'] = stats['OK']
            worksheet[f'C{row}'] = stats['WARNING']
            worksheet[f'D{row}'] = stats['CRITICAL']
            row += 1
    
    # Create the chart
    chart = BarChart()
    chart.title = "Security Status by Component"
    chart.x_axis.title = "Components"
    chart.y_axis.title = "Number of domains"
    
    data = Reference(worksheet, min_col=2, min_row=start_row, max_col=4, max_row=row-1)
    categories = Reference(worksheet, min_col=1, min_row=start_row+1, max_row=row-1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    
    # Style
    chart.type = "col"
    chart.style = 10    
    # Positioning
    chart.anchor = f'F{start_row}'
    chart.width = 20
    chart.height = 12
    
    worksheet.add_chart(chart)

# ----------------- Main avec rapport final -----------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="🔍 ULTRA-DETAILED EMAIL AUDITOR",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
📋 This script analyzes your email configurations in depth and explains:
   • WHY each element is critical
   • WHAT HAPPENS in case of misconfiguration
   • HOW to fix detected problems
   • WHAT IS THE BUSINESS IMPACT of each vulnerability

Usage examples:
   python3 checkdmarc_enhanced.py scan1.json scan2.json
   python3 checkdmarc_enhanced.py /path/to/checkdmarc/output/*.json -excel
        """
    )
    
    parser.add_argument('json_files', nargs='+', 
                       help='JSON files generated by checkdmarc to analyze')
    parser.add_argument('-excel', '--excel', action='store_true',
                       help='Generate a complete Excel report with charts')
    
    args = parser.parse_args()
    
    if not args.json_files:
        parser.print_help()
        sys.exit(1)
    
    # Automatically detect output directory
    output_dir = None
    if args.excel:
        # Use the directory of the first JSON file as base
        first_file = Path(args.json_files[0])
        if first_file.exists():
            output_dir = first_file.parent
        else:
            output_dir = Path.cwd()
        
        print(f"📂 Excel output directory: {output_dir}")

    overall_ok = True
    total_domains = 0
    critical_issues = 0
    warning_issues = 0
    
    # If Excel mode only, silent processing
    if args.excel:
        print("📊 Excel generation mode activated - Silent analysis processing...")
        
        # Generate Excel report
        generate_excel_report(args.json_files, output_dir)
        return
    
    # Normal mode with detailed display
    for file in args.json_files:
        path = Path(file)
        if not path.exists():
            print(f"❌ File not found: {file}")
            continue

        total_domains += 1
        print(f"\n🎯 ===== EMAIL SECURITY AUDIT FOR: {path.stem.upper()} =====")
        
        try:
            data = json.loads(path.read_text())
            statuses = audit_domain(data)
            
            domain_critical = 0
            domain_warnings = 0
            
            for st in statuses:
                print_status(st)
                if st[0] == "CRITICAL":
                    overall_ok = False
                    domain_critical += 1
                    critical_issues += 1
                elif st[0] == "WARNING":
                    domain_warnings += 1
                    warning_issues += 1
            
            # Domain summary
            print("=" * 60)
            print(f"📊 SUMMARY FOR {path.stem.upper()}")
            print("=" * 60)
            if domain_critical == 0 and domain_warnings == 0:
                print("🏆 EXCELLENT! Exemplary email configuration!")
            elif domain_critical == 0:
                print(f"✅ GOOD! {domain_warnings} recommended improvements")
            else:
                print(f"🚨 CRITICAL! {domain_critical} major issues + {domain_warnings} warnings")
            print()
            
        except Exception as e:
            print(f"❌ Error analyzing {file}: {e}")
            overall_ok = False

    # Global final report
    print("\n" + "=" * 80)
    print("🎯 FINAL REPORT - EMAIL SECURITY AUDIT")
    print("=" * 80)
    
    if total_domains == 1:
        if overall_ok:
            print("🏆 CONGRATULATIONS! Your domain is properly secured.")
        else:
            print("⚠️ WARNING! Critical vulnerabilities have been detected.")
    else:
        print(f"📊 DOMAINS ANALYZED: {total_domains}")
        print(f"🚨 CRITICAL ISSUES: {critical_issues}")
        print(f"⚠️ RECOMMENDED IMPROVEMENTS: {warning_issues}")
        
        if overall_ok:
            print("🏆 OVERALL RESULT: All your domains are properly protected!")
        else:
            print("⚠️ OVERALL RESULT: Immediate corrective actions are required.")
    
    print("\n💡 RECOMMENDED NEXT STEPS:")
    if critical_issues > 0:
        print("   1. 🚨 URGENT: Fix CRITICAL issues IMMEDIATELY")
        print("   2. ⚠️ Plan improvements for warnings")
        print("   3. 🔄 Re-scan after corrections")
    elif warning_issues > 0:
        print("   1. ⚠️ Plan recommended improvements")
        print("   2. 🔄 Re-scan after optimizations")
        print("   3. 📊 Monitor DMARC reports regularly")
    else:
        print("   1. 📊 Monitor DMARC reports monthly")
        print("   2. 🔄 Re-scan quarterly")
        print("   3. 🏆 Maintain security excellence!")
    
    print(f"\n⏰ AUDIT COMPLETED - Return code: {'0 (success)' if overall_ok else '1 (issues detected)'}")
    sys.exit(0 if overall_ok else 1)

if __name__ == "__main__":
    main()
